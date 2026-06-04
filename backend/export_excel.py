"""
会計データ確認シート Excel出力モジュール
openpyxl + VML Form Control チェックボックスで全Excelバージョン対応。

- F列（対応済）: Form Control チェックボックス（1クリックで ✓/□）
- チェックを入れると行全体がグレーアウト（条件付き書式）
- Excel 2010 以降すべてのバージョンで動作
"""
import io, re, zipfile, textwrap
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── カラー定数 ──────────────────────────────────────────
C_ORANGE_BG   = "F47A20"
C_ORANGE_PALE = "FDE8D8"
C_ORANGE_HEAD = "7A3800"
C_WHITE       = "FFFFFF"
C_ROW_ALT     = "FDF9F5"
C_DONE_BG     = "D0D0D0"
C_DONE_FG     = "888888"
C_BORDER      = "E8D5C0"
C_INSTR_BG    = "FEF5ED"
C_INSTR_FG    = "9A7060"

THIN = Side(style="thin",   color=C_BORDER)
MED  = Side(style="medium", color="C09060")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MED = Border(left=MED,  right=MED,  top=MED,  bottom=MED)


def _fmt(bold=False, size=10, fg=None, bg=None,
         halign="left", valign="center", wrap=False, border=None):
    f = Font(name="メイリオ", size=size, bold=bold, color=fg or "000000")
    al = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    fill = PatternFill("solid", fgColor=bg) if bg else PatternFill()
    kwargs = dict(font=f, alignment=al, fill=fill)
    if border:
        kwargs["border"] = border
    # openpyxlでは直接 NamedStyle でまとめられないので辞書で返す
    return kwargs


def _apply(cell, *, bold=False, size=10, fg=None, bg=None,
           halign="left", valign="center", wrap=False, border=BORDER_ALL):
    cell.font      = Font(name="メイリオ", size=size, bold=bold, color=fg or "000000")
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.border = border


# ────────────────────────────────────────────────────────
# メイン関数
# ────────────────────────────────────────────────────────
def build_checksheet_xlsx(issues: list, period: str = "", client_name: str = "") -> bytes:
    """
    チェック結果を xlsx に変換して bytes で返す。
    F列に Form Control チェックボックスを配置し、
    クリックすると行全体がグレーアウトする。
    """
    wb = Workbook()
    ws_check  = wb.active
    ws_check.title = "チェックシート"
    ws_sample = wb.create_sheet("記入例")

    data_start_row = 5     # Excel行番号（1始まり）
    n_issues       = len(issues)

    _build_sheet(wb, ws_check, issues, period, data_start_row)
    _build_sample_sheet(wb, ws_sample)

    # --- 一旦 bytes に保存 ---
    raw_buf = io.BytesIO()
    wb.save(raw_buf)
    raw_bytes = raw_buf.getvalue()

    # --- zip を開いて VML チェックボックスを注入 ---
    final_bytes = _inject_checkboxes(
        raw_bytes,
        sheet_name="チェックシート",
        data_start_row=data_start_row,
        n_rows=n_issues,
        f_col_index=6,        # F列 = 6列目（1始まり）
        linked_col_index=7,   # G列をリンクセルとして使用（非表示）
    )
    return final_bytes


# ────────────────────────────────────────────────────────
# チェックシート本体
# ────────────────────────────────────────────────────────
def _build_sheet(wb, ws, issues, period, data_start_row):
    # 列幅
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 0.1   # リンクセル（非表示）

    # 行高
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 22

    # Row1: タイトル
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "会計データ確認シート"
    _apply(c, bold=True, size=14, fg=C_WHITE, bg=C_ORANGE_BG,
           halign="center", border=BORDER_MED)

    # Row2: 対象月
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"対象月：{period or '　　　　年　　　月分'}　　　　　　　　　　　　　　担当："
    _apply(c, size=11, bg="FFFAF6", border=BORDER_ALL)

    # Row3: 操作方法
    ws.merge_cells("A3:F3")
    c = ws["A3"]
    c.value = "【操作方法】　F列（対応済）のチェックボックスをクリック → ✓ でグレーアウト、もう一度クリックで解除"
    _apply(c, size=9, fg=C_INSTR_FG, bg=C_INSTR_BG, border=BORDER_ALL)

    # Row4: ヘッダー
    for col_idx, h in enumerate(
        ["勘定科目", "補助科目", "日付", "確認内容（Harel入力欄）", "ご回答", "対応済"],
        start=1
    ):
        c = ws.cell(row=4, column=col_idx, value=h)
        _apply(c, bold=True, size=11, fg=C_ORANGE_HEAD, bg=C_ORANGE_PALE,
               halign="center", border=BORDER_ALL)

    # 条件付き書式: G列（リンクセル）がTRUEのとき行をグレーアウト
    last_row = max(data_start_row + len(issues) + 5, 50)
    gray_fill = PatternFill("solid", fgColor=C_DONE_BG)
    gray_font = Font(name="メイリオ", color=C_DONE_FG)
    ds   = DifferentialStyle(fill=gray_fill, font=gray_font)
    rule = Rule(type="expression", dxf=ds,
                formula=[f"$G{data_start_row}=TRUE"])
    ws.conditional_formatting.add(
        f"A{data_start_row}:F{last_row}", rule
    )

    # データ行
    prev_base = None
    for i, issue in enumerate(issues):
        row = data_start_row + i

        full_acc = str(issue.get("account", ""))
        month    = str(issue.get("month",   ""))
        message  = str(issue.get("message", ""))

        if "（" in full_acc and full_acc.endswith("）"):
            p    = full_acc.index("（")
            base = full_acc[:p]
            sub  = full_acc[p+1:-1]
        else:
            base = full_acc
            sub  = ""

        display_base = base if base != prev_base else ""
        if base:
            prev_base = base

        date_str = month
        if len(month) == 7 and month[4] == "-":
            y, m = month.split("-")
            date_str = f"{y}年{int(m)}月"

        clean_msg = re.sub(r'^【[^】]+】', '', message).strip()
        is_alt    = (i % 2 == 1)
        row_bg    = C_ROW_ALT if is_alt else C_WHITE

        values = [display_base, sub, date_str, clean_msg, ""]
        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            _apply(c, size=10, bg=row_bg,
                   halign="center" if col_idx == 3 else "left",
                   wrap=(col_idx == 4))

        # F列は空（チェックボックスが重なる）
        c = ws.cell(row=row, column=6, value="")
        _apply(c, size=10, bg=row_bg, halign="center")

        # G列: リンクセル（非表示・初期値FALSE）
        ws.cell(row=row, column=7, value=False)

        # 行高
        lines = max(1, len(clean_msg) // 40 + 1)
        ws.row_dimensions[row].height = max(18, min(lines * 15, 80))

    # G列を非表示に
    ws.column_dimensions["G"].hidden = True


# ────────────────────────────────────────────────────────
# 記入例シート
# ────────────────────────────────────────────────────────
def _build_sample_sheet(wb, ws):
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 10

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "会計データ確認シート"
    _apply(c, bold=True, size=14, fg=C_WHITE, bg=C_ORANGE_BG,
           halign="center", border=BORDER_MED)

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = "対象月：　　　　年　　　月分　　　　　　　　　　　　　　　担当："
    _apply(c, size=11)

    for col_idx, h in enumerate(
        ["勘定科目", "補助科目", "日付", "確認内容（Harel入力欄）", "ご回答", "対応済"],
        start=1
    ):
        c = ws.cell(row=3, column=col_idx, value=h)
        _apply(c, bold=True, size=11, fg=C_ORANGE_HEAD, bg=C_ORANGE_PALE,
               halign="center", border=BORDER_ALL)

    samples = [
        ("売掛金", "ｱｲﾑﾋﾞｰ",  "", "3月分158,130円の回収がないようです"),
        ("",      "ｻﾛﾝｻｻﾞﾝ", "", "4/30の売り上げが2件入っているのはないか"),
        ("",      "諸口",    "", "前期から-1,320円残っているので確認が必要"),
        ("買掛金", "箔一産業", "", "３月分の仕入計上がされていないか　4/30 2,337,775円支払分"),
        ("",      "ﾌﾟﾗﾑ",    "", "3月計上 260,271円の支払いがない　翌々月以降支払いか"),
        ("未払費用","役員退職金","", "永井等さま分 30,000,000円が残っている。支払日未達で問題ないか"),
    ]
    for i, (acc, sub, date, content) in enumerate(samples):
        row = 4 + i
        bg  = C_ROW_ALT if i % 2 == 1 else C_WHITE
        for col, val in enumerate([acc, sub, date, content, "", ""], start=1):
            c = ws.cell(row=row, column=col, value=val)
            _apply(c, size=10, bg=bg,
                   halign="center" if col == 3 else "left",
                   wrap=(col == 4))
        ws.row_dimensions[row].height = 20


# ────────────────────────────────────────────────────────
# VML Form Control チェックボックス注入
# ────────────────────────────────────────────────────────
def _inject_checkboxes(xlsx_bytes: bytes, sheet_name: str,
                       data_start_row: int, n_rows: int,
                       f_col_index: int, linked_col_index: int) -> bytes:
    """
    xlsx の zip を開き、VML チェックボックスを追加して返す。
    各データ行の F 列に Form Control チェックボックスを配置し、
    G 列（非表示）をリンクセルとして使用する。
    """
    if n_rows == 0:
        return xlsx_bytes

    # Zip を読み込む
    in_zip  = io.BytesIO(xlsx_bytes)
    out_zip = io.BytesIO()

    with zipfile.ZipFile(in_zip, "r") as zin:
        names = zin.namelist()

        # シート番号を特定
        sheet_idx = None
        for n in names:
            if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"):
                content = zin.read(n).decode("utf-8")
                if sheet_name in content or True:  # 簡易: sheet1を使う
                    num = n.replace("xl/worksheets/sheet", "").replace(".xml", "")
                    if num.isdigit():
                        sheet_idx = int(num)
                        break

        if sheet_idx is None:
            sheet_idx = 1

        sheet_xml_path = f"xl/worksheets/sheet{sheet_idx}.xml"
        vml_path       = f"xl/drawings/vmlDrawing{sheet_idx}.vml"
        rel_path       = f"xl/worksheets/_rels/sheet{sheet_idx}.xml.rels"
        ct_path        = "[Content_Types].xml"

        # VML を生成
        vml_content = _make_vml(data_start_row, n_rows,
                                f_col_index, linked_col_index)

        with zipfile.ZipFile(out_zip, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in names:
                data = zin.read(item)

                if item == sheet_xml_path:
                    # シートXMLに legacy drawing 参照を追加
                    data = _add_legacy_drawing(data, sheet_idx)

                elif item == rel_path:
                    # リレーションシップに vmlDrawing を追加
                    data = _add_vml_rel(data, sheet_idx)

                elif item == ct_path:
                    # Content Types に vml を追加
                    data = _add_content_type(data)

                zout.writestr(item, data)

            # VML ファイルを新規追加
            zout.writestr(vml_path, vml_content)

            # rels ファイルが存在しない場合は新規作成
            if rel_path not in names:
                zout.writestr(rel_path, _make_rels(sheet_idx))

    return out_zip.getvalue()


def _make_vml(data_start_row: int, n_rows: int,
              f_col: int, linked_col: int) -> str:
    """VML チェックボックス XML を生成する"""
    # f_col, linked_col は 1始まり
    f_col0      = f_col - 1        # 0始まりのExcel列インデックス
    linked_col0 = linked_col - 1

    shapes = []
    shape_id_base = 1025

    for i in range(n_rows):
        row0    = data_start_row - 1 + i   # 0始まり行インデックス
        row1    = data_start_row + i       # 1始まり行番号（Excelの行番号）
        shape_id = shape_id_base + i
        linked_cell = f"${get_column_letter(linked_col)}${row1}"

        # Anchor: leftCol, leftDX, topRow, topDY, rightCol, rightDX, bottomRow, bottomDY
        # F列内に収まるよう配置（列インデックスは0始まり）
        left_col  = f_col0
        right_col = f_col0
        top_row   = row0
        bot_row   = row0

        shapes.append(f"""  <v:shape id="_x0000_s{shape_id}" type="#_x0000_t201"
    style="position:absolute;margin-left:2pt;margin-top:2pt;width:68pt;height:14pt;z-index:{i+1}"
    o:insetmode="auto">
    <v:fill color2="window [65]"/>
    <v:shadow on="t" color="black" obscured="t"/>
    <v:path o:connecttype="none"/>
    <x:ClientData ObjectType="Checkbox">
      <x:Anchor>{left_col}, 36, {top_row}, 4, {right_col}, 220, {bot_row}, 14</x:Anchor>
      <x:PrintObject/>
      <x:AutoFill>False</x:AutoFill>
      <x:FmlaCB/>
      <x:FmlaLink>{linked_cell}</x:FmlaLink>
      <x:TextHAlign>Left</x:TextHAlign>
      <x:NoThreeD/>
    </x:ClientData>
  </v:shape>""")

    shapes_xml = "\n".join(shapes)

    return f"""<xml xmlns:v="urn:schemas-microsoft-com:vml"
     xmlns:o="urn:schemas-microsoft-com:office:office"
     xmlns:x="urn:schemas-microsoft-com:office:excel">
  <o:shapelayout v:ext="edit">
    <o:idmap v:ext="edit" data="1"/>
  </o:shapelayout>
  <v:shapetype id="_x0000_t201" coordsize="21600,21600" o:spt="201"
    path="m,l,21600r21600,xe">
    <v:stroke joinstyle="miter"/>
    <v:path shadowok="f" o:extrusionok="f" gradientshapeok="t" o:connecttype="rect"/>
  </v:shapetype>
{shapes_xml}
</xml>"""


def _add_legacy_drawing(sheet_xml: bytes, sheet_idx: int) -> bytes:
    """シートXMLに <legacyDrawing> タグを追加する"""
    xml_str = sheet_xml.decode("utf-8")
    tag = f'<legacyDrawing r:id="rIdVml{sheet_idx}"/>'
    if "legacyDrawing" in xml_str:
        return sheet_xml
    # </worksheet> の直前に挿入
    xml_str = xml_str.replace("</worksheet>", f"{tag}</worksheet>")
    return xml_str.encode("utf-8")


def _add_vml_rel(rels_xml: bytes, sheet_idx: int) -> bytes:
    """rels ファイルに VML リレーションシップを追加する"""
    xml_str = rels_xml.decode("utf-8")
    rel = (f'<Relationship Id="rIdVml{sheet_idx}" '
           f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
           f'Target="../drawings/vmlDrawing{sheet_idx}.vml"/>')
    if "vmlDrawing" in xml_str:
        return rels_xml
    xml_str = xml_str.replace("</Relationships>", f"{rel}</Relationships>")
    return xml_str.encode("utf-8")


def _add_content_type(ct_xml: bytes) -> bytes:
    """Content_Types.xml に vml の Override を追加する"""
    xml_str = ct_xml.decode("utf-8")
    if "vmlDrawing" in xml_str:
        return ct_xml
    override = ('<Default Extension="vml" '
                'ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>')
    xml_str = xml_str.replace("</Types>", f"{override}</Types>")
    return xml_str.encode("utf-8")


def _make_rels(sheet_idx: int) -> str:
    """シートのリレーションシップファイルを新規作成する"""
    rel = (f'<Relationship Id="rIdVml{sheet_idx}" '
           f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" '
           f'Target="../drawings/vmlDrawing{sheet_idx}.vml"/>')
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        f'{rel}'
        '</Relationships>'
    )
